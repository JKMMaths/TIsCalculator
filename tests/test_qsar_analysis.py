import io
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from src.qsar_analysis import (
    add_topological_index_columns,
    correlation_table,
    create_all_property_graphs_zip,
    create_correlation_plot,
    create_qsar_report,
    run_qsar_analysis,
)


def test_regression_analysis_and_report():
    smiles = ["C" * length for length in range(1, 13)] * 2
    data = pd.DataFrame({"SMILES": smiles, "Solubility": [float(len(value)) for value in smiles]})

    result = run_qsar_analysis(data, "SMILES", "Solubility", "Regression", test_size=0.25)

    assert result.task_type == "regression"
    assert len(result.metrics) == 4
    assert set(result.metrics["Model"]) == {
        "Classical ML (Random Forest)", "Deep Learning (MLP)",
        "Graph Theory (Random Forest)", "Hybrid AI (RF, all features)",
    }
    assert not result.predictions.empty

    workbook = load_workbook(io.BytesIO(create_qsar_report(result)), read_only=True)
    assert workbook.sheetnames == ["Run Summary", "Model Metrics", "Test Predictions", "Excluded Rows"]


def test_invalid_rows_are_reported():
    smiles = ["C" * length for length in range(1, 13)] + ["not-smiles"]
    data = pd.DataFrame({"SMILES": smiles, "Target": list(range(13))})

    result = run_qsar_analysis(data, "SMILES", "Target", "Regression")

    assert len(result.excluded_rows) == 1
    assert result.excluded_rows.iloc[0]["Reason"] == "Invalid SMILES"


def test_classification_analysis():
    smiles = ["C" * length for length in range(1, 13)] * 2
    labels = ["low" if len(value) <= 6 else "high" for value in smiles]
    data = pd.DataFrame({"SMILES": smiles, "Activity": labels})

    result = run_qsar_analysis(data, "SMILES", "Activity", "Classification", test_size=0.25)

    assert result.task_type == "classification"
    assert {"Accuracy", "Balanced Accuracy", "F1", "ROC AUC"}.issubset(result.metrics.columns)
    assert set(result.predictions["Actual"]).issubset({"low", "high"})


def test_ti_property_correlation_graph_downloads():
    smiles = ["C" * length for length in range(1, 13)]
    data = pd.DataFrame({"SMILES": smiles, "Boiling point": [20.0 + 7.5 * len(value) for value in smiles]})
    augmented = add_topological_index_columns(data, "SMILES")
    ti_columns = [column for column in augmented if column.startswith("TI | ")]

    correlations = correlation_table(augmented, "Boiling point", ti_columns)
    assert not correlations.empty
    assert correlations.iloc[0]["Sample Count"] == 12

    png = create_correlation_plot(augmented, "Boiling point", correlations.iloc[0]["TI Column"])
    assert png.startswith(b"\x89PNG")

    archive_bytes = create_all_property_graphs_zip(augmented, ["Boiling point"], ti_columns)
    with ZipFile(io.BytesIO(archive_bytes)) as archive:
        assert "correlation_summary.csv" in archive.namelist()
        assert any(name.endswith(".png") for name in archive.namelist())
