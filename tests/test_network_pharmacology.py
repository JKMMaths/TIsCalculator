import io
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from src.network_pharmacology import (
    NetworkPharmacologyResult, create_cytoscape_zip, create_report,
    gene_table, intersect_targets, normalize_genes, rank_hubs, venn_figure,
)


def test_normalization_intersection_and_provenance():
    assert normalize_genes("akt1, EGFR; akt1\nTNF") == ["AKT1", "EGFR", "TNF"]
    drug = gene_table("AKT1 EGFR TNF", "SwissTargetPrediction", "Drug target")
    disease = gene_table("IL6, TNF, egfr", "DisGeNET", "Disease association")
    overlap = intersect_targets(drug, disease)
    assert overlap["Gene"].tolist() == ["EGFR", "TNF"]
    assert set(drug["Source"]) == {"SwissTargetPrediction"}


def test_hubs_include_isolated_intersection_genes():
    edges = pd.DataFrame({
        "Protein A": ["AKT1", "AKT1", "EGFR"], "Protein B": ["EGFR", "TNF", "TNF"],
        "STRING ID A": ["a", "a", "b"], "STRING ID B": ["b", "c", "c"],
        "Combined score": [0.9, 0.8, 0.7],
    })
    hubs = rank_hubs(["AKT1", "EGFR", "TNF", "IL6"], edges)
    assert set(hubs["Gene"]) == {"AKT1", "EGFR", "TNF", "IL6"}
    assert hubs.iloc[-1]["Gene"] == "IL6"
    assert hubs.iloc[-1]["Degree"] == 0


def test_report_visual_and_cytoscape_exports():
    drug = gene_table("AKT1 TNF", "Source A", "Drug target")
    disease = gene_table("TNF IL6", "Source B", "Disease association")
    overlap = intersect_targets(drug, disease)
    edges = pd.DataFrame(columns=["Protein A", "Protein B", "STRING ID A", "STRING ID B", "Combined score"])
    hubs = rank_hubs(overlap["Gene"], edges)
    result = NetworkPharmacologyResult("Compound", "Disease", 9606, drug, disease, overlap, edges, hubs, pd.DataFrame())

    assert venn_figure(2, 2, 1).startswith(b"\x89PNG")
    workbook = load_workbook(io.BytesIO(create_report(result)), read_only=True)
    assert {"Summary", "Drug Targets", "Disease Genes", "Intersection", "Hub Proteins"}.issubset(workbook.sheetnames)
    with ZipFile(io.BytesIO(create_cytoscape_zip(result))) as archive:
        assert {"nodes.csv", "edges.csv", "enrichment.csv", "README.txt"}.issubset(archive.namelist())
